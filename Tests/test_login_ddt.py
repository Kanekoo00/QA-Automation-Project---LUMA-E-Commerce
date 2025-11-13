import pytest
import allure
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from Pages.Test_login_page import LoginPage


def get_login_data():
    """Load login test data from Excel"""
    workbook = load_workbook("Data/login_data.xlsx")
    sheet = workbook.active
    data = []
    for row in range(2, sheet.max_row + 1):
        email = sheet.cell(row, 1).value
        password = sheet.cell(row, 2).value
        expected = sheet.cell(row, 3).value
        data.append((email, password, expected))
    return data


@allure.title("Login Test (Data Driven)")
@pytest.mark.parametrize("email, password, expected", get_login_data())
def test_login_data_driven(email, password, expected):
    driver = webdriver.Chrome()
    driver.maximize_window()
    login_page = LoginPage(driver)

    try:
        with allure.step("Open login page"):
            login_page.open()
            login_page.close_cookie_popup()

        with allure.step(f"Enter credentials: {email} / {password}"):
            login_page.enter_email(email)
            login_page.enter_password(password)
            login_page.click_login()

        with allure.step("Validate login result"):
            current_url = driver.current_url
            error_message = ""

            # Tunggu 3–5 detik untuk pesan error jika login gagal
            try:
                error_element = WebDriverWait(driver, 5).until(
                    EC.visibility_of_element_located(
                        (By.XPATH, "//*[contains(text(), 'User name and password do not match')]")
                        )
                    )
                error_message = error_element.text.strip()
            except:
                pass  # tidak ada pesan error berarti mungkin login sukses

            #  Validasi hasil
            if expected == "success":
                assert "/en.html" in current_url, f"Expected success but got URL: {current_url}"
                print(f"Login success — redirected to homepage: {current_url}")
            else:
                assert "User name and password do not match" in error_message, \
                    f"Expected failure message not found. Got: '{error_message}'"
                print(f"Login failed as expected: {error_message}")

    finally:
        driver.quit()
